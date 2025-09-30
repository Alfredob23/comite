import os
from django.shortcuts import render,redirect
from .models import Ingresos,Usuarios,Egresos,Facturar,Biologicos,Mensajes
from django.http import JsonResponse,HttpResponse
from django.template.loader import get_template
from io import BytesIO
from django.template.loader import render_to_string
import io
import openpyxl
from openpyxl.styles import Font
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.contrib import messages
from django.db.models import Sum

def home(request):
    return render(request,"home.html")



@login_required
def dashboard(request):
    total_ingresos = Ingresos.objects.count()
    total_egresos = Egresos.objects.count()
    total_facturas = Facturar.objects.count()
    suma_ingresos = int(Ingresos.objects.aggregate(total=Sum('valorIngreso'))['total'] or 0)
    suma_egresos = int(Egresos.objects.aggregate(total=Sum('valorEgreso'))['total'] or 0)
    suma_facturas = int(Facturar.objects.aggregate(total=Sum('valor_total'))['total'] or 0)
    return render(request, "dashboard.html", {
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "total_facturas": total_facturas,
        "suma_ingresos": suma_ingresos,
        "suma_egresos": suma_egresos,
        "suma_facturas": suma_facturas,
    })
 
# INGRESOS
@login_required
def ingresos(request):
    ingresos = Ingresos.objects.all()
    usuarios = Usuarios.objects.all()
    ingresos_invertidos = []
    for i in ingresos[::-1]:
        ingresos_invertidos.append(i)
    return render(request,"ingresos/ingresos.html",{"ingresos":ingresos_invertidos,"usuarios":usuarios})

@login_required
def registrarIngreso(request):
    factura_ingreso = request.POST['factura']
    cedula = request.POST['cedula']
    nombre_completo = request.POST['nombre_completo']
    direccion = request.POST['direccion']
    tipo_pago = request.POST['tipo_pago']
    valorIngreso = request.POST['valorIngreso']
    concepto = request.POST['concepto']
    factura = ''
    if factura_ingreso:
        factura = Facturar.objects.get(nFactura=factura_ingreso)
    if cedula and nombre_completo and direccion:
        usuario, created = Usuarios.objects.get_or_create(cedula=cedula, defaults={'nombre_completo': nombre_completo,'direccion': direccion})
                # Si el usuario fue creado correctamente, creamos el ingreso
        if usuario  and factura:
            Ingresos.objects.create(usuario=usuario, tipo_pago=tipo_pago,valorIngreso=valorIngreso,concepto=concepto,facturas=factura)
            return redirect('/ingresos')
        else:
            Ingresos.objects.create(usuario=usuario, tipo_pago=tipo_pago,valorIngreso=valorIngreso,concepto=concepto)
            return redirect('/ingresos')
    else:
        return render(request, '/', {'error': 'Todos los campos son requeridos.'})
    return render(request, '/')

@login_required
def eliminacionIngreso(request,nIngreso):
    ingreso = Ingresos.objects.get(nIngreso=nIngreso)
    ingreso.facturas
    ingreso.delete()
    return redirect('/ingresos')

@login_required
def edicionIngreso(request,nIngreso):
    ingreso = Ingresos.objects.get(nIngreso=nIngreso)
    
    return render(request,"ingresos/edicionIngresos.html",{"ingreso":ingreso})


@login_required
def detalleIngreso(request,nIngreso):
    ingreso = Ingresos.objects.get(nIngreso=nIngreso)
    return render(request,"ingresos/detalleIngresos.html",{"ingreso":ingreso})

@login_required
def editarIngreso(request,nIngreso):
    pago = request.POST['tipo_pago']
    valor =request.POST['valorIngreso']
    concepto =request.POST['concepto']   
    ingreso = Ingresos.objects.get(nIngreso=nIngreso)
    ingreso.tipo_pago = pago
    ingreso.valorIngreso = int(valor)
    ingreso.concepto = concepto
    ingreso.save()
    return redirect('/')

@login_required
def verificar_cedula(request):
    cedula = request.GET.get('cedula', None)
    if cedula:
        try:
            usuario = Usuarios.objects.get(cedula=cedula)
            data = {
                'usuario': {
                    'nombre_completo': usuario.nombre_completo,
                    'direccion': usuario.direccion,
                }
            }
        except Usuarios.DoesNotExist:
            data = {'usuario': None}
    else:
        data = {'usuario': None}
    
    return JsonResponse(data)


@login_required
def detallePDF(request,nIngreso):
    ingreso = Ingresos.objects.get(nIngreso=nIngreso)
    return render(request,"ingresos/ticketIngresos.html",{"ingreso":ingreso})


@login_required
def export_to_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Personas'
        
    sheet.append(['nIngreso', 'Nombre', 'Cedula','Direccion', 'Metodo de Pago', 'Valor','Concepto', 'Ciudad', 'Fecha'])
    for cell in sheet[1]:
        cell.font = Font(bold=True) 
    
    ingresos = Ingresos.objects.all()
    
    for ingreso in ingresos:
        if ingreso.fecha:
            fecha = ingreso.fecha.replace(tzinfo=None)
        else:
            fecha= None
        sheet.append([ingreso.nIngreso, ingreso.usuario.nombre_completo, ingreso.usuario.cedula,ingreso.usuario.direccion, ingreso.tipo_pago, ingreso.valorIngreso,ingreso.concepto, ingreso.ciudad, fecha])
         
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)  # Encuentra el texto más largo
        adjusted_width = max_length + 2  # Añadir un pequeño margen
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ingresos.xlsx"'     

    workbook.save(response)

    return response


######EGRESOS######


@login_required
def egresos(request):
    egresos = Egresos.objects.all()
    egresos_invertidos = []
    for i in egresos[::-1]:
        egresos_invertidos.append(i)
    return render(request,"egresos/egresos.html",{"egresos":egresos_invertidos})

@login_required
def registrarEgreso(request):
    tipo_pago = request.POST['tipo_pago']
    valorEgreso = request.POST['valorEgreso']
    if request.POST['concepto'] == 'Otro':
        concepto = request.POST['otroConcepto']
    else:
        concepto = request.POST['concepto']
    recibe = request.POST['recibe']
    revisado = request.POST['revisado']
    egreso = Egresos.objects.get_or_create(tipo_pago=tipo_pago,valorEgreso = valorEgreso,concepto = concepto,recibe = recibe,revisado = revisado)
    return redirect('/egresos')


@login_required
def edicionEgreso(request,nEgreso):
    egreso = Egresos.objects.get(nEgreso=nEgreso)
    return render(request,"egresos/edicionEgresos.html",{"egreso":egreso})


@login_required
def eliminacionEgreso(request,nEgreso):
    egreso = Egresos.objects.get(nEgreso=nEgreso)
    egreso.delete()
    return redirect('/egresos')


@login_required
def detalleEgreso(request,nEgreso):
    egreso = Egresos.objects.get(nEgreso=nEgreso)
    return render(request,"egresos/detalleEgreso.html",{"egreso":egreso})


@login_required
def editarEgreso(request,nEgreso):
    tipo_pago = request.POST['tipo_pago']
    valorEgreso = request.POST['valorEgreso']
    if request.POST['concepto'] == 'Otro':
        concepto = request.POST['otroConcepto']
    else:
        concepto = request.POST['concepto']
    recibe = request.POST['recibe']
    revisado = request.POST['revisado']
    egreso = Egresos.objects.get(nEgreso=nEgreso)
    egreso.tipo_pago = tipo_pago
    egreso.valorEgreso = valorEgreso
    egreso.concepto = concepto
    egreso.recibe = recibe
    egreso.revisado = revisado
    egreso.save()    
    return redirect('/egresos')


@login_required
def imprimirEgreso(request,nEgreso):
    egreso = Egresos.objects.get(nEgreso=nEgreso)
    return render(request,"egresos/ticketEgresos.html",{"egreso":egreso})

@login_required
def export_to_excel_egreso(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Personas'
    
        
    sheet.append(['No Egreso', 'Tipo Pago', 'Valor','Concepto', 'Ciudad', 'Recibe','Aprobado', 'Revisado','Contabilizado','Fecha'])
    for cell in sheet[1]:
        cell.font = Font(bold=True) 
    
    egresos = Egresos.objects.all()
    
    for egreso in egresos:
        if egreso.fecha:
            fecha = egreso.fecha.replace(tzinfo=None)
        else:
            fecha= None
        sheet.append([egreso.nEgreso, egreso.tipo_pago, egreso.valorEgreso,egreso.concepto,egreso.ciudad,egreso.recibe,egreso.aprobado,egreso.revisado,egreso.contabilizado,fecha])
        
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)  # Encuentra el texto más largo
        adjusted_width = max_length + 2  # Añadir un pequeño margen
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="egresos.xlsx"'     

    workbook.save(response)

    return response

#####FACTURAR#####


@login_required
def facturar(request):
    factura = Facturar.objects.all()
    precios_biologicos = {biologico.nombre: biologico.valorUnidad for biologico in Biologicos.objects.all()}
    factura_invertida = []
    for i in factura[::-1]:
        factura_invertida.append(i)
    return render(request,"facturar/facturas.html",{"facturas":factura_invertida,"precios_biologicos": precios_biologicos})


@login_required
def registrarFactura(request):
    cedula = request.POST['cedula']
    nombre_completo = request.POST['nombre_completo']
    direccion = request.POST['direccion']
    biologico = request.POST.getlist('biologico[]')
    lote = request.POST.getlist('lote[]')
    cantidad = request.POST.getlist('cantidad[]')
    cantidad_int = [int(elemento) for elemento in cantidad]
    cantidad_aftosa = 0
    cantidad_cepa19 = 0
    laboratorio = []
    
    #En esta parte recorro los diferentes numeros de lote y de acuerdo al rango se encuentre
    #le asigno un nombre.
    for l,b in zip(lote,biologico):  
        if int(l) in range(200,351) and b =='Aftosa':
            laboratorio.append('Limor')
        elif int(l) in range(400,601) and b =='Aftosa':
            laboratorio.append('Vecol')
        elif int(l) > 800 and b =='Cepa19':
            laboratorio.append('Vecol')
                 
    for b,c in zip (biologico,cantidad):     
        if b == 'Aftosa':
            cantidad_aftosa +=int(c)
        elif b == 'Cepa19':
            cantidad_cepa19 +=int(c)
    if cedula and nombre_completo and direccion and biologico:
            usuario, created = Usuarios.objects.get_or_create(cedula=cedula, defaults={'nombre_completo': nombre_completo,'direccion': direccion})
                    # Si el usuario fue creado correctamente, creamos el ingreso
            if usuario:
                Facturar.objects.create(usuario=usuario, cantidad_aftosa =cantidad_aftosa,cantidad_cepa19 =cantidad_cepa19,cantidad_total=cantidad_int,lote = lote,biologico = biologico,laboratorio=laboratorio)
                return redirect('/facturar/')
    else:
        return render(request, '/', {'error': 'Todos los campos son requeridos.'})
    return render(request, '/facturar/')


@login_required
def obtener_precio_biologico(request):
    # Obtener el nombre del biológico desde la solicitud AJAX
    nombre_biologico = request.GET.get('biologico', '')

    # Buscar el precio del biológico en el modelo
    try:
        biologico = Biologicos.objects.get(nombre=nombre_biologico)
        precio = biologico.valorUnidad
        return JsonResponse({'precio': precio})
    except Biologicos.DoesNotExist:
        return JsonResponse({'error': 'Biológico no encontrado'}, status=404)


@login_required
def edicionFactura(request,nFactura):
    factura = Facturar.objects.get(nFactura = nFactura)
    biologicos = list(factura.biologico)
    cantidad_total = list(factura.cantidad_total)
    lote = list(factura.lote)
    factura_pares = zip(biologicos, cantidad_total, lote)
    context = {
    "factura_pares": factura_pares,"factura":factura,
    }
    return render(request, 'facturar/edicionFactura.html',context)


@login_required
def editarFactura(request,nFactura):
    cedula = request.POST['cedula']
    nombre_completo = request.POST['nombre_completo']
    direccion = request.POST['direccion']
    biologico = request.POST.getlist('biologico[]')
    lote = request.POST.getlist('lote[]')
    cantidad = request.POST.getlist('cantidad[]')
    cantidad_int = [int(elemento) for elemento in cantidad]
    cantidad_aftosa = 0
    cantidad_cepa19 = 0
    laboratorio = []
    
    #En esta parte recorro los diferentes numeros de lote y de acuerdo al rango se encuentre
    #le asigno un nombre del laborario.
    for l,b in zip(lote,biologico):  
        if int(l) in range(200,351) and b =='Aftosa':
            laboratorio.append('Limor')
        elif int(l) in range(400,601) and b =='Aftosa':
            laboratorio.append('Vecol')
        elif int(l) > 800 and b =='Cepa19':
            laboratorio.append('Vecol')
        else:
            laboratorio.append('null')
                 
    for b,c in zip (biologico,cantidad):     
        if b == 'Aftosa':
            cantidad_aftosa +=int(c)
        elif b == 'Cepa19':
            cantidad_cepa19 +=int(c)

    factura = Facturar.objects.get(nFactura=nFactura)
    factura.cedula = cedula
    factura.nombre_completo = nombre_completo
    factura.direccion = direccion
    factura.biologico = biologico
    factura.lote = lote
    factura.cantidad_total = cantidad_int
    factura.cantidad_aftosa = cantidad_aftosa
    factura.cantidad_cepa19 = cantidad_cepa19
    factura.laboratorio = laboratorio
    factura.save()    
    return redirect('/facturar')


@login_required
def eliminacionFactura(request,nFactura):
    factura = Facturar.objects.get(nFactura=nFactura)
    factura.delete()
    return redirect('/facturar')


@login_required
def detalleFactura(request,nFactura):
    factura = Facturar.objects.get(nFactura=nFactura)
    ingresos = Ingresos.objects.filter(facturas__nFactura=nFactura)
    print(ingresos)
    return render(request, "facturar/detalleFactura.html", {
        "factura": factura,
        "ingresos": ingresos
    })


@login_required
def imprimirFactura(request,nFactura):
    factura = Facturar.objects.get(nFactura=nFactura)
    return render(request,"facturar/ticketFactura.html",{"factura":factura})


@login_required
def export_to_excel_facturas(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Facturas'
    
        
    sheet.append(['No Factura', 'Nombre', 'Cedula','C. Aftosa', 'C. Cepa19', 'Valor total','Fecha'])
    for cell in sheet[1]:
        cell.font = Font(bold=True) 
    
    
    
    facturas = Facturar.objects.all()
    
    for factura in facturas:
        if factura.fecha:
            fecha = factura.fecha.replace(tzinfo=None)
        else:
            fecha= None
        sheet.append([factura.nFactura, factura.usuario.nombre_completo, factura.usuario.cedula,factura.cantidad_aftosa ,factura.cantidad_cepa19 ,factura.valor_total,fecha])
        
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value) 
        adjusted_width = max_length + 2  
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Facturas.xlsx"'     

    workbook.save(response)

    return response

@login_required
def get_factura_data(request):
    if request.method == 'GET':
        factura_number = request.GET.get('factura', None)
        if factura_number:
            try:
                factura = Facturar.objects.get(nFactura=factura_number)
                response_data = {
                    'cedula': factura.usuario.cedula, 
                    'nombre_completo': factura.usuario.nombre_completo,  
                    'direccion': factura.usuario.direccion,
                    'valorIngreso': factura.valor_pendiente,
                }
                return JsonResponse({'success': True, 'data': response_data})
            except Facturar.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Factura no encontrada'})
        return JsonResponse({'success': False, 'error': 'Número de factura no proporcionado'})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

###MENSAJES#####

def mensajes(request):
    nombre = request.POST['nombre']
    correo = request.POST['correo']
    telefono= request.POST['telefono']
    mensaje = request.POST['mensaje']
    Mensajes.objects.create(nombre=nombre, correo =correo,telefono =telefono,mensaje=mensaje)
    return redirect('/')
@login_required
def ver_mensajes(request):
    mensajes = Mensajes.objects.all().order_by('-id')
    return render(request, "mensajes/mensajes.html", {
        "mensajes": mensajes,
    })
@login_required
def detalle_mensaje(request,id):
    mensaje = Mensajes.objects.get(id=id)
    if not mensaje.leido:
        mensaje.leido = True
        mensaje.save()
    return render(request,"mensajes/detalleMensaje.html",{"mensaje":mensaje})
@login_required
def eliminar_mensaje(request,id):
    mensaje = Mensajes.objects.get(id=id)
    mensaje.delete()
    return redirect("/bandeja-mensajes/")

@login_required
def responder_mensaje(request,id):
    mensaje = Mensajes.objects.get(id=id)
    return render(request,"mensajes/responderMensaje.html",{"mensaje":mensaje})

@login_required
def enviar_correo(request):
    if request.method == 'POST':
        correo = request.POST.get('correo')
        asunto = request.POST.get('asunto')
        mensaje = request.POST.get('mensaje')

        try:
            send_mail(
                subject=asunto,
                message=mensaje,
                from_email=None,
                recipient_list=[correo],
                fail_silently=False,
            )

            messages.success(request, 'Correo enviado con éxito')
        except Exception as e:
            messages.error(request, f'Error al enviar el correo: {str(e)}')

        return render(request, "mensajes/responderMensaje.html", {
            "mensaje": {
                "correo": correo,
                "asunto": asunto,
            }
        })

    return render(request, "mensajes/mensajes.html")

